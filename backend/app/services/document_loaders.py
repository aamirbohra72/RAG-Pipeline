"""
Document loaders for Markdown, HTML, and structured data (CSV/XLSX).

Each loader returns blocks compatible with chunk_text_blocks and populates
doc_type + section metadata for pgvector rows.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import List, Tuple

_HEADER = re.compile(r"^(#{1,6})\s+(.+)$", re.MULTILINE)


@dataclass
class LoaderResult:
    blocks: List[Tuple[int, str, str]] = field(default_factory=list)
    doc_type: str = ""
    meta: dict = field(default_factory=dict)


def _doc_type_from_filename(filename: str) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    mapping = {
        "md": "markdown",
        "markdown": "markdown",
        "html": "html",
        "htm": "html",
        "csv": "csv",
        "xlsx": "xlsx",
        "pdf": "pdf",
    }
    return mapping.get(ext, ext or "unknown")


def supported_extensions() -> set[str]:
    return {".pdf", ".md", ".markdown", ".html", ".htm", ".csv", ".xlsx"}


class _TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._parts: list[str] = []
        self._skip = False

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag in {"script", "style", "noscript"}:
            self._skip = True
        elif tag in {"p", "br", "div", "li", "h1", "h2", "h3", "h4", "h5", "h6", "tr"}:
            self._parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style", "noscript"}:
            self._skip = False
        elif tag in {"p", "div", "li", "tr", "table"}:
            self._parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self._skip and data.strip():
            self._parts.append(data.strip() + " ")

    def text(self) -> str:
        return re.sub(r"\n{3,}", "\n\n", "".join(self._parts)).strip()


def load_markdown(file_bytes: bytes, filename: str) -> LoaderResult:
    text = file_bytes.decode("utf-8", errors="replace")
    doc_type = "markdown"
    blocks: List[Tuple[int, str, str]] = []

    sections: list[tuple[str, list[str]]] = [("document", [])]
    current_section = "document"

    for line in text.splitlines():
        header_match = _HEADER.match(line)
        if header_match:
            current_section = header_match.group(2).strip()
            sections.append((current_section, []))
        else:
            sections[-1][1].append(line)

    page_num = 1
    for section_name, lines in sections:
        body = "\n".join(lines).strip()
        if body:
            blocks.append((page_num, section_name, body))
            page_num += 1

    if not blocks and text.strip():
        blocks.append((1, "document", text.strip()))

    return LoaderResult(blocks=blocks, doc_type=doc_type, meta={"filename": filename})


def load_html(file_bytes: bytes, filename: str) -> LoaderResult:
    raw = file_bytes.decode("utf-8", errors="replace")
    parser = _TextExtractor()
    parser.feed(raw)
    text = parser.text()
    if not text.strip():
        raise ValueError(f"No extractable text in HTML file {filename}")

    # Split on blank lines into pseudo-sections
    blocks: List[Tuple[int, str, str]] = []
    page_num = 1
    for para in re.split(r"\n{2,}", text):
        cleaned = para.strip()
        if cleaned:
            blocks.append((page_num, f"section {page_num}", cleaned))
            page_num += 1

    return LoaderResult(blocks=blocks, doc_type="html", meta={"filename": filename})


def _row_template(headers: list[str], row: dict) -> str:
    """Preserve column context per chunk — not a bare CSV serialization."""
    lines = [f"Record from structured data ({', '.join(headers)}):"]
    for key in headers:
        val = row.get(key, "")
        if val is not None and str(val).strip():
            lines.append(f"  {key}: {val}")
    return "\n".join(lines)


def load_csv(file_bytes: bytes, filename: str) -> LoaderResult:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError(f"CSV {filename} has no headers")

    headers = list(reader.fieldnames)
    blocks: List[Tuple[int, str, str]] = []
    for idx, row in enumerate(reader, start=1):
        content = _row_template(headers, row)
        section = f"row {idx}"
        blocks.append((idx, section, content))

    if not blocks:
        raise ValueError(f"CSV {filename} has no data rows")

    return LoaderResult(
        blocks=blocks,
        doc_type="csv",
        meta={"filename": filename, "columns": headers, "rows": len(blocks)},
    )


def load_xlsx(file_bytes: bytes, filename: str) -> LoaderResult:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    blocks: List[Tuple[int, str, str]] = []
    page_num = 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
        if not any(headers):
            continue

        row_idx = 0
        for values in rows:
            row_idx += 1
            row = {headers[i]: (values[i] if i < len(values) else "") for i in range(len(headers))}
            if not any(v is not None and str(v).strip() for v in row.values()):
                continue
            content = _row_template(headers, row)
            section = f"{sheet_name} / row {row_idx}"
            blocks.append((page_num, section, content))
            page_num += 1

    wb.close()
    if not blocks:
        raise ValueError(f"XLSX {filename} has no data rows")

    return LoaderResult(blocks=blocks, doc_type="xlsx", meta={"filename": filename, "rows": len(blocks)})


def load_document(file_bytes: bytes, filename: str) -> LoaderResult:
    ext = Path(filename).suffix.lower()
    if ext in {".md", ".markdown"}:
        return load_markdown(file_bytes, filename)
    if ext in {".html", ".htm"}:
        return load_html(file_bytes, filename)
    if ext == ".csv":
        return load_csv(file_bytes, filename)
    if ext == ".xlsx":
        return load_xlsx(file_bytes, filename)
    raise ValueError(f"Unsupported file type: {filename}")
